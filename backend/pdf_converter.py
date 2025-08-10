#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF変換機能（クロスプラットフォーム対応）
ExcelファイルをPDFに変換する - Windows/Mac対応
"""

import os
import glob
import platform
import subprocess
import sys
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple, Union, TYPE_CHECKING
from pathlib import Path
import shutil
import tempfile

# Excel読み込み用
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl not available")
    OPENPYXL_AVAILABLE = False
    # フォールバック用のダミー定義
    class Worksheet:
        pass

# PDF作成用
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.pdfbase import pdfutils
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    REPORTLAB_AVAILABLE = True
except ImportError:
    print("Warning: reportlab not available")
    REPORTLAB_AVAILABLE = False

# Excel制御用（macOS/Windows両対応だがここでは主にmacOSで使用）
try:
    import xlwings as xw  # type: ignore
    XLWINGS_AVAILABLE = True
except Exception:
    XLWINGS_AVAILABLE = False

# 日本語フォント対応
JAPANESE_FONT = 'Helvetica'
if REPORTLAB_AVAILABLE:
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        JAPANESE_FONT = 'HeiseiMin-W3'
    except:
        # フォントが利用できない場合はHelveticaを使用
        JAPANESE_FONT = 'Helvetica'


class PDFConverter:
    """クロスプラットフォーム対応PDF変換クラス"""
    
    def __init__(self):
        self.platform = platform.system()
        self._check_dependencies()
        
    def _check_dependencies(self) -> Dict[str, bool]:
        """依存関係の確認"""
        deps = {
            'openpyxl': OPENPYXL_AVAILABLE,
            'reportlab': REPORTLAB_AVAILABLE,
            'platform': True
        }
        # WindowsでExcelネイティブ出力を使う場合に必要
        if self.platform == 'Windows':
            try:
                import win32com.client  # type: ignore
                import pythoncom  # type: ignore
                deps['pywin32'] = True
            except Exception:
                deps['pywin32'] = False
        return deps
    
    def get_system_info(self) -> Dict[str, Any]:
        """システム情報を取得"""
        deps = self._check_dependencies()
        return {
            'platform': self.platform,
            'python_version': sys.version,
            'dependencies': deps,
            'missing_deps': [k for k, v in deps.items() if not v]
        }

    def _is_command_available(self, command: str) -> bool:
        try:
            result = shutil.which(command)
            return result is not None
        except Exception:
            return False

    def _is_macos_excel_available(self) -> bool:
        if self.platform != 'Darwin':
            return False
        excel_app_path = '/Applications/Microsoft Excel.app'
        return os.path.exists(excel_app_path) and self._is_command_available('osascript')

    def _is_macos_xlwings_available(self) -> bool:
        return self.platform == 'Darwin' and XLWINGS_AVAILABLE

    def _is_windows_excel_available(self) -> bool:
        if self.platform != 'Windows':
            return False
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
        except Exception:
            return False
        excel_app = None
        try:
            pythoncom.CoInitialize()
            try:
                excel_app = win32com.client.Dispatch("Excel.Application")
                return True
            except Exception:
                return False
        finally:
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        
    def get_excel_files(self, folder_path: str) -> Dict[str, Any]:
        """
        指定フォルダ内のExcelファイルを取得
        
        Args:
            folder_path: フォルダパス
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            folder_path = os.path.normpath(folder_path)
            
            if not os.path.exists(folder_path):
                return {
                    'success': False,
                    'error': f'フォルダが見つかりません: {folder_path}',
                    'error_type': 'folder_not_found'
                }
            
            if not os.path.isdir(folder_path):
                return {
                    'success': False,
                    'error': f'指定されたパスはフォルダではありません: {folder_path}',
                    'error_type': 'not_directory'
                }
            
            # Excelファイルを検索
            excel_patterns = ['*.xlsx', '*.xls', '*.xlsm']
            excel_files = []
            
            for pattern in excel_patterns:
                try:
                    files = glob.glob(os.path.join(folder_path, pattern))
                    excel_files.extend(files)
                except Exception as e:
                    print(f"Warning: Failed to search pattern {pattern}: {str(e)}")
                    continue
            
            # ファイル情報を取得
            file_list = []
            failed_files = []
            
            for file_path in excel_files:
                try:
                    # macOS向けにパスを正規化
                    try:
                        if sys.platform == 'darwin':
                            file_path = str(Path(file_path))
                    except Exception:
                        pass
                    file_name = os.path.basename(file_path)
                    file_size = os.path.getsize(file_path)
                    
                    # ファイルの読み取り権限を確認
                    if not os.access(file_path, os.R_OK):
                        failed_files.append({
                            'path': file_path,
                            'error': 'ファイルの読み取り権限がありません'
                        })
                        continue
                    
                    file_list.append({
                        'path': file_path,
                        'name': file_name,
                        'size': file_size
                    })
                except Exception as e:
                    failed_files.append({
                        'path': file_path,
                        'error': f"ファイル情報取得エラー: {str(e)}"
                    })
                    continue
            
            result = {
                'success': True,
                'files': file_list,
                'count': len(file_list),
                'folder_path': folder_path
            }
            
            if failed_files:
                result['failed_files'] = failed_files
                result['warning'] = f'{len(failed_files)}個のファイルでエラーが発生しました'
            
            return result
            
        except Exception as e:
            return {
                'success': False,
                'error': f'ファイル取得エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def create_output_folder(self, base_output_dir: str) -> Dict[str, Any]:
        """
        出力フォルダを作成（日付フォルダ）
        
        Args:
            base_output_dir: 基本出力ディレクトリ
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            base_output_dir = os.path.normpath(base_output_dir)
            
            # 親ディレクトリの存在確認
            if not os.path.exists(base_output_dir):
                try:
                    Path(base_output_dir).mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    return {
                        'success': False,
                        'error': f'ベースディレクトリの作成に失敗: {str(e)}',
                        'error_type': 'base_dir_creation_failed'
                    }
            
            # 現在の日付でフォルダ名を生成（YYYYMM形式）
            current_date = datetime.now()
            folder_name = current_date.strftime("%Y%m")
            
            # 出力フォルダパスを作成
            output_folder = os.path.join(base_output_dir, folder_name)
            
            # フォルダを作成
            try:
                Path(output_folder).mkdir(parents=True, exist_ok=True)
            except PermissionError:
                return {
                    'success': False,
                    'error': f'フォルダの作成権限がありません: {output_folder}',
                    'error_type': 'permission_denied'
                }
            except Exception as e:
                return {
                    'success': False,
                    'error': f'フォルダ作成エラー: {str(e)}',
                    'error_type': 'folder_creation_failed'
                }
            
            # 書き込み権限の確認
            if not os.access(output_folder, os.W_OK):
                return {
                    'success': False,
                    'error': f'出力フォルダに書き込み権限がありません: {output_folder}',
                    'error_type': 'write_permission_denied'
                }
            
            return {
                'success': True,
                'output_folder': output_folder,
                'folder_name': folder_name,
                'base_dir': base_output_dir
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'出力フォルダ作成エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def _validate_excel_file(self, file_path: str) -> Tuple[bool, str]:
        """Excelファイルの妥当性を検証"""
        try:
            if not OPENPYXL_AVAILABLE:
                return False, "openpyxl が利用できません"
            
            if not os.path.exists(file_path):
                return False, "ファイルが見つかりません"
            
            if not os.access(file_path, os.R_OK):
                return False, "ファイルの読み取り権限がありません"
            
            # ファイルサイズの確認
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                return False, "ファイルが空です"
            
            if file_size > 100 * 1024 * 1024:  # 100MB制限
                return False, "ファイルサイズが大きすぎます (100MB以上)"
            
            # Excelファイルの読み込みテスト
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                if not workbook.sheetnames:
                    return False, "シートが見つかりません"
                workbook.close()
            except Exception as e:
                return False, f"Excelファイルの読み込みエラー: {str(e)}"
            
            return True, "OK"
            
        except Exception as e:
            return False, f"ファイル検証エラー: {str(e)}"
    
    def _excel_to_pdf_openpyxl(self, excel_file: str, pdf_path: str) -> Tuple[bool, str]:
        """
        openpyxlを使用してExcelからPDFを生成
        
        Args:
            excel_file: Excelファイルのパス
            pdf_path: 出力PDFのパス
            
        Returns:
            (成功/失敗, エラーメッセージ)
        """
        try:
            if not REPORTLAB_AVAILABLE:
                return False, "reportlab が利用できません"
            
            # Excelファイルを読み込み（macOS ではパスを正規化）
            try:
                if sys.platform == 'darwin':
                    excel_file = str(Path(excel_file))
            except Exception:
                pass
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # PDFドキュメントを作成
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=20*mm,
                bottomMargin=20*mm
            )
            
            story = []
            styles = getSampleStyleSheet()
            # 日本語対応（段落ヘッダーは作らず、シンプルに表のみ出力）
            normal_jp = ParagraphStyle(
                'NormalJP',
                parent=styles['Normal'],
                fontName=JAPANESE_FONT,
                fontSize=9
            )
            
            # FMT（独自見出し）は出力しない
            
            # 各シートを処理
            processed_sheets = 0
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    
                    # シートのデータを取得
                    data = self._get_sheet_data(sheet)
                    
                    if data:
                        # テーブルを作成
                        table = Table(data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('FONTNAME', (0, 1), (-1, -1), JAPANESE_FONT),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        story.append(table)
                        processed_sheets += 1
                    else:
                        story.append(Paragraph("", normal_jp))
                    
                    story.append(PageBreak())
                    
                except Exception as e:
                    # ログのみ。PDFには余計な文言は出さない
                    story.append(PageBreak())
                    continue
            
            workbook.close()
            
            if processed_sheets == 0:
                return False, "処理可能なシートが見つかりませんでした"
            
            # PDFを生成
            doc.build(story)
            
            # ファイルが正常に作成されたかチェック
            if not os.path.exists(pdf_path):
                return False, "PDFファイルの作成に失敗しました"
            
            if os.path.getsize(pdf_path) == 0:
                return False, "PDFファイルが空です"
            
            return True, f"成功 ({processed_sheets} シート処理)"
            
        except Exception as e:
            return False, f"PDF変換エラー: {str(e)}"

    def _excel_to_pdf_win32(self, excel_files: List[str], output_folder: str) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """
        Windows + Excel(COM)でワークブック内の全シートをPDF化（高速・レイアウト忠実）

        Args:
            excel_files: 変換するExcelファイルパス一覧
            output_folder: 出力フォルダ

        Returns:
            (converted_files, failed_files) のタプル
        """
        converted_files: List[Dict[str, str]] = []
        failed_files: List[Dict[str, str]] = []
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
            from win32com.client import constants  # type: ignore
        except Exception as e:
            return [], [{
                'file': '(batch)',
                'error': f'pywin32がインストールされていません: {str(e)}'
            }]

        excel_app = None
        try:
            pythoncom.CoInitialize()
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.ScreenUpdating = False
            excel_app.DisplayAlerts = False
            try:
                # マクロやセキュリティ関連のダイアログ抑止
                # msoAutomationSecurityForceDisable = 3
                excel_app.AutomationSecurity = 3  # type: ignore[attr-defined]
            except Exception:
                pass
            try:
                # リンク更新やイベントによるダイアログ抑止
                excel_app.AskToUpdateLinks = False  # type: ignore[attr-defined]
            except Exception:
                pass
            try:
                excel_app.EnableEvents = False  # type: ignore[attr-defined]
            except Exception:
                pass

            for excel_file in excel_files:
                try:
                    # PDF出力先パスを作成
                    base_name = os.path.splitext(os.path.basename(excel_file))[0]
                    pdf_name = f"{base_name}.pdf"
                    pdf_path = os.path.join(output_folder, pdf_name)
                    if os.path.exists(pdf_path):
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        pdf_path = os.path.join(output_folder, f"{base_name}_{timestamp}.pdf")

                    # ブックを開く（読み取り専用・リンク更新や推奨読み取り専用を無視）
                    wb = excel_app.Workbooks.Open(
                        excel_file,
                        ReadOnly=True,
                        UpdateLinks=False,
                        IgnoreReadOnlyRecommended=True,
                        Editable=False,
                        Notify=False,
                        AddToMru=False,
                        Local=True,
                    )
                    try:
                        # ブック単位でPDFへ（全シート対象・印刷設定を尊重）
                        # Type=0 (xlTypePDF)
                        wb.ExportAsFixedFormat(
                            0,
                            pdf_path,
                            Quality=0,  # xlQualityStandard
                            IncludeDocProperties=True,
                            IgnorePrintAreas=False,  # 既存の印刷範囲/ページ設定を尊重
                            OpenAfterPublish=False
                        )
                    finally:
                        wb.Close(SaveChanges=False)

                    if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                        converted_files.append({
                            'excel_file': excel_file,
                            'pdf_file': pdf_path,
                            'pdf_name': os.path.basename(pdf_path),
                            'message': 'Excelの全シートをPDFとして保存'
                        })
                    else:
                        failed_files.append({
                            'file': excel_file,
                            'error': 'PDFファイルが作成されませんでした'
                        })
                except Exception as e:
                    failed_files.append({
                        'file': excel_file,
                        'error': f'Excel出力エラー: {str(e)}'
                    })

        except Exception as e:
            failed_files.append({
                'file': '(batch)',
                'error': f'Excelアプリ起動エラー: {str(e)}'
            })
        finally:
            try:
                if excel_app is not None:
                    excel_app.DisplayAlerts = True
                    excel_app.ScreenUpdating = True
                    excel_app.Quit()
            except Exception:
                pass
            try:
                import pythoncom  # type: ignore
                pythoncom.CoUninitialize()
            except Exception:
                pass

        return converted_files, failed_files

    def _find_python_executable(self) -> Optional[str]:
        """実行可能なPythonコマンドを探す（Windows用フォールバックに利用）"""
        candidates: List[str] = []
        try:
            if sys.executable and ('python' in os.path.basename(sys.executable).lower()):
                candidates.append(sys.executable)
        except Exception:
            pass
        # 環境変数に指定があれば優先
        if os.environ.get('PYTHON'):
            candidates.append(os.environ['PYTHON'])
        # 一般的な候補
        candidates.extend(['py', 'python', 'python3'])

        for c in candidates:
            try:
                proc = subprocess.run([c, '--version'], capture_output=True, text=True, timeout=5)
                if proc.returncode == 0:
                    return c
            except Exception:
                continue
        return None

    def _excel_to_pdf_win32_one_subprocess(self, excel_file: str, output_folder: str, timeout_seconds: int = 90) -> Tuple[bool, str, Optional[str]]:
        """
        単一Excel→PDF変換をサブプロセスで実行（ハング耐性向上）

        Returns: (ok, error_message, pdf_path)
        """
        python_cmd = self._find_python_executable()
        if not python_cmd:
            return False, 'Python実行環境が見つかりません（サブプロセス方式）', None

        child_code = r'''
import sys, os
try:
    import pythoncom
    import win32com.client
    from datetime import datetime
except Exception as e:
    print(str(e), file=sys.stderr)
    sys.exit(1)

excel_file = sys.argv[1]
output_folder = sys.argv[2]
pythoncom.CoInitialize()
excel_app = None
try:
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False
    excel_app.ScreenUpdating = False
    excel_app.DisplayAlerts = False
    try:
        excel_app.AutomationSecurity = 3
    except Exception:
        pass
    try:
        excel_app.AskToUpdateLinks = False
    except Exception:
        pass
    try:
        excel_app.EnableEvents = False
    except Exception:
        pass

    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    pdf_path = os.path.join(output_folder, base_name + ".pdf")
    if os.path.exists(pdf_path):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_path = os.path.join(output_folder, f"{base_name}_{ts}.pdf")

    wb = excel_app.Workbooks.Open(
        excel_file,
        ReadOnly=True,
        UpdateLinks=False,
        IgnoreReadOnlyRecommended=True,
        Editable=False,
        Notify=False,
        AddToMru=False,
        Local=True,
    )
    try:
        wb.ExportAsFixedFormat(
            0,
            pdf_path,
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
    finally:
        wb.Close(SaveChanges=False)

    if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
        print('PDFファイルの作成に失敗しました', file=sys.stderr)
        sys.exit(2)

    print(pdf_path)
    sys.exit(0)
except Exception as e:
    import traceback
    print(str(e), file=sys.stderr)
    print(traceback.format_exc(), file=sys.stderr)
    sys.exit(1)
finally:
    try:
        if excel_app is not None:
            excel_app.DisplayAlerts = True
            excel_app.ScreenUpdating = True
            excel_app.Quit()
    except Exception:
        pass
    try:
        pythoncom.CoUninitialize()
    except Exception:
        pass
'''

        try:
            proc = subprocess.run(
                [python_cmd, '-c', child_code, excel_file, output_folder],
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
            )
            if proc.returncode == 0:
                pdf_path = proc.stdout.strip().splitlines()[-1] if proc.stdout else None
                return True, '', pdf_path
            return False, (proc.stderr or 'サブプロセス変換エラー'), None
        except subprocess.TimeoutExpired:
            return False, f'サブプロセスがタイムアウトしました（{timeout_seconds}秒）', None
        except Exception as e:
            return False, str(e), None

    def _excel_to_pdf_macos_excel(self, excel_files: List[str], output_folder: str, timeout_seconds: int = 90) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """macOSのMicrosoft ExcelをAppleScript経由で用いてPDF出力"""
        converted_files: List[Dict[str, str]] = []
        failed_files: List[Dict[str, str]] = []

        if not self._is_macos_excel_available():
            return [], [{'file': '(batch)', 'error': 'Microsoft Excel (macOS) または osascript が見つかりません'}]

        # AppleScript（引数: 入力/出力）
        applescript = r'''
on run argv
    set inputPathPosix to item 1 of argv
    set outputPathPosix to item 2 of argv
    tell application "Microsoft Excel"
        activate
        try
            set display alerts to false
        end try
        set wb to open (POSIX file inputPathPosix) read only yes
        try
            -- ActiveWorkbook 全体を印刷範囲に従ってPDF化
            set vbCmd to "ActiveWorkbook.ExportAsFixedFormat Type:=0, Filename:=\"" & outputPathPosix & "\", Quality:=0, IncludeDocProperties:=True, IgnorePrintAreas:=False, OpenAfterPublish:=False"
            do visual basic vbCmd
        on error errMsg number errNum
            try
                close wb saving no
            end try
            error errMsg number errNum
        end try
        close wb saving no
    end tell
end run
'''

        # 一時ファイルに保存
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.scpt', delete=False) as f:
                script_path = f.name
                f.write(applescript)
        except Exception as e:
            return [], [{'file': '(batch)', 'error': f'AppleScript作成エラー: {str(e)}'}]

        try:
            for excel_file in excel_files:
                try:
                    base_name = os.path.splitext(os.path.basename(excel_file))[0]
                    pdf_path = os.path.join(output_folder, f"{base_name}.pdf")
                    if os.path.exists(pdf_path):
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        pdf_path = os.path.join(output_folder, f"{base_name}_{timestamp}.pdf")

                    proc = subprocess.run([
                        'osascript', script_path, excel_file, pdf_path
                    ], capture_output=True, text=True, timeout=timeout_seconds)

                    if proc.returncode == 0 and os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                        converted_files.append({
                            'excel_file': excel_file,
                            'pdf_file': pdf_path,
                            'pdf_name': os.path.basename(pdf_path),
                            'message': 'Excel (macOS) によるPDF保存'
                        })
                    else:
                        err = proc.stderr.strip() or 'Excel (macOS) 変換エラー'
                        failed_files.append({'file': excel_file, 'error': err})
                except subprocess.TimeoutExpired:
                    failed_files.append({'file': excel_file, 'error': f'Excel (macOS) がタイムアウトしました（{timeout_seconds}秒）'})
                except Exception as e:
                    failed_files.append({'file': excel_file, 'error': str(e)})
        finally:
            try:
                os.remove(script_path)
            except Exception:
                pass

        return converted_files, failed_files

    def _excel_to_pdf_macos_xlwings(self, excel_files: List[str], output_folder: str) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """macOSのMicrosoft Excelをxlwings経由で用いてPDF出力（印刷範囲尊重・全シート）"""
        converted_files: List[Dict[str, str]] = []
        failed_files: List[Dict[str, str]] = []
        if not XLWINGS_AVAILABLE:
            return converted_files, [{'file': '(batch)', 'error': 'xlwingsが利用できません'}]
        app = None
        try:
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False
            for excel_file in excel_files:
                try:
                    base_name = os.path.splitext(os.path.basename(excel_file))[0]
                    pdf_path = os.path.join(output_folder, f"{base_name}.pdf")
                    if os.path.exists(pdf_path):
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        pdf_path = os.path.join(output_folder, f"{base_name}_{timestamp}.pdf")

                    wb = app.books.open(excel_file, update_links=False, read_only=True)
                    try:
                        # すべてのシートを対象にPDF出力（xlwings標準API）
                        # macOSでは内部的にAppleScriptを利用
                        wb.to_pdf(path=pdf_path)
                    finally:
                        # xlwingsのBook.closeは引数なしが正しい（macOSでSaveChangesキーワードは未対応）
                        wb.close()

                    if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                        converted_files.append({
                            'excel_file': excel_file,
                            'pdf_file': pdf_path,
                            'pdf_name': os.path.basename(pdf_path),
                            'message': 'Excel (xlwings) によるPDF保存'
                        })
                    else:
                        failed_files.append({'file': excel_file, 'error': 'PDFファイルが作成されませんでした'})
                except Exception as e:
                    failed_files.append({'file': excel_file, 'error': str(e)})
        except Exception as e:
            failed_files.append({'file': '(batch)', 'error': f'xlwingsでのExcel起動エラー: {str(e)}'})
        finally:
            try:
                if app is not None:
                    app.display_alerts = True
                    app.screen_updating = True
                    app.quit()
            except Exception:
                pass
        return converted_files, failed_files
    
    def _get_sheet_data(self, sheet: Any, max_rows: int = 100, max_cols: int = 20) -> List[List[str]]:
        """
        シートからデータを取得
        
        Args:
            sheet: ワークシート
            max_rows: 最大行数
            max_cols: 最大列数
            
        Returns:
            データリスト
        """
        try:
            data: List[List[str]] = []

            # 使用範囲を決定
            try:
                max_row_attr = getattr(sheet, 'max_row', 0) or 0
                max_col_attr = getattr(sheet, 'max_column', 0) or 0
                max_row = min(max_row_attr if isinstance(max_row_attr, int) else 0, max_rows)
                max_col = min(max_col_attr if isinstance(max_col_attr, int) else 0, max_cols)
            except Exception:
                max_row = min(100, max_rows)
                max_col = min(20, max_cols)

            if max_row <= 0 or max_col <= 0:
                return []

            # 高速な一括取得（values_only=True）
            iter_rows = getattr(sheet, 'iter_rows', None)
            if callable(iter_rows):
                for values in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                    row_output: List[str] = []
                    has_text = False
                    for value in values:
                        if value is None:
                            s = ''
                        elif isinstance(value, (int, float)):
                            s = str(value)
                        else:
                            s = str(value)
                            if len(s) > 100:
                                s = s[:100] + "..."
                        if isinstance(s, str) and s.strip():
                            has_text = True
                        row_output.append(s)
                    if has_text:
                        data.append(row_output)
            else:
                # フォールバック（古いAPIやWorksheet互換）
                for row in range(1, max_row + 1):
                    row_output = []
                    has_text = False
                    for col in range(1, max_col + 1):
                        try:
                            cell = sheet.cell(row=row, column=col) if hasattr(sheet, 'cell') else None
                            value = getattr(cell, 'value', None)
                            if value is None:
                                s = ''
                            elif isinstance(value, (int, float)):
                                s = str(value)
                            else:
                                s = str(value)
                                if len(s) > 100:
                                    s = s[:100] + "..."
                            if isinstance(s, str) and s.strip():
                                has_text = True
                            row_output.append(s)
                        except Exception:
                            row_output.append('')
                    if has_text:
                        data.append(row_output)

            return data

        except Exception as e:
            print(f"シートデータ取得エラー: {str(e)}")
            return []
    
    def convert_to_pdf(self, excel_files: List[str], output_folder: str) -> Dict[str, Any]:
        """
        ExcelファイルをPDFに変換（クロスプラットフォーム対応）
        
        Args:
            excel_files: 変換するExcelファイルのパスリスト
            output_folder: 出力フォルダパス
            
        Returns:
            結果辞書
        """
        try:
            # 出力フォルダの存在確認（なければ作成）
            try:
                Path(output_folder).mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            if not os.path.exists(output_folder):
                return {
                    'success': False,
                    'error': f'出力フォルダが見つかりません: {output_folder}',
                    'error_type': 'output_folder_not_found'
                }

            converted_files: List[Dict[str, str]] = []
            failed_files: List[Dict[str, str]] = []
            validation_errors: List[Dict[str, str]] = []

            # プラットフォーム別処理
            if self.platform == 'Windows':
                # Windows: Excel(デスクトップ版)必須。未インストール時はエラー
                if not self._is_windows_excel_available():
                    return {
                        'success': False,
                        'error': 'Excelがインストールされていません',
                        'error_type': 'excel_not_installed'
                    }
                conv, fail = self._excel_to_pdf_win32(excel_files, output_folder)
                converted_files.extend(conv)
                failed_files.extend(fail)

            elif self.platform == 'Darwin':
                # macOS: まずExcel(デスクトップ版)経由（xlwings優先→AppleScript）を試み、失敗時はopenpyxl+reportlabでフォールバック
                if self._is_macos_xlwings_available():
                    conv, fail = self._excel_to_pdf_macos_xlwings(excel_files, output_folder)
                    converted_files.extend(conv)
                    failed_files.extend(fail)
                elif self._is_macos_excel_available():
                    conv, fail = self._excel_to_pdf_macos_excel(excel_files, output_folder)
                    converted_files.extend(conv)
                    failed_files.extend(fail)
                else:
                    # Excelが無い場合は、reportlabがあればフォールバックを試す
                    if OPENPYXL_AVAILABLE and REPORTLAB_AVAILABLE:
                        for excel_path in excel_files:
                            base_name = os.path.splitext(os.path.basename(excel_path))[0]
                            pdf_path = os.path.join(output_folder, f"{base_name}.pdf")
                            ok, msg = self._excel_to_pdf_openpyxl(excel_path, pdf_path)
                            if ok:
                                converted_files.append({
                                    'excel_file': excel_path,
                                    'pdf_file': pdf_path,
                                    'pdf_name': os.path.basename(pdf_path),
                                    'message': 'openpyxl+reportlab によるPDF保存'
                                })
                            else:
                                failed_files.append({'file': excel_path, 'error': msg})
                    # すべて失敗した場合のフォールバック
                    if len(converted_files) == 0 and len(failed_files) > 0 and OPENPYXL_AVAILABLE and REPORTLAB_AVAILABLE:
                        for excel_path in excel_files:
                            base_name = os.path.splitext(os.path.basename(excel_path))[0]
                            pdf_path = os.path.join(output_folder, f"{base_name}.pdf")
                            ok, msg = self._excel_to_pdf_openpyxl(excel_path, pdf_path)
                            if ok:
                                converted_files.append({
                                    'excel_file': excel_path,
                                    'pdf_file': pdf_path,
                                    'pdf_name': os.path.basename(pdf_path),
                                    'message': 'openpyxl+reportlab フォールバックPDF保存'
                                })
                            else:
                                # 既に失敗に入っている場合は重複させない
                                if not any(f.get('file') == excel_path for f in failed_files):
                                    failed_files.append({'file': excel_path, 'error': msg})
            else:
                # その他プラットフォームは非対応
                return {
                    'success': False,
                    'error': 'Excelがインストールされていません',
                    'error_type': 'excel_not_installed'
                }

            # 結果の整理
            result: Dict[str, Any] = {
                'success': True,
                'converted_files': converted_files,
                'failed_files': failed_files,
                'validation_errors': validation_errors,
                'total_converted': len(converted_files),
                'total_failed': len(failed_files),
                'total_validation_errors': len(validation_errors),
                'output_folder': output_folder
            }

            if len(converted_files) == 0 and (len(failed_files) > 0 or len(validation_errors) > 0):
                result['success'] = False
                result['error'] = 'すべてのファイルの変換に失敗しました'

            return result

        except Exception as e:
            return {
                'success': False,
                'error': f'PDF変換エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def open_folder(self, folder_path: str) -> Dict[str, Any]:
        """
        フォルダを開く（クロスプラットフォーム対応）
        
        Args:
            folder_path: 開くフォルダのパス
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            folder_path = os.path.normpath(folder_path)
            
            if not os.path.exists(folder_path):
                return {
                    'success': False,
                    'error': f'フォルダが見つかりません: {folder_path}',
                    'error_type': 'folder_not_found'
                }
            
            if not os.path.isdir(folder_path):
                return {
                    'success': False,
                    'error': f'指定されたパスはフォルダではありません: {folder_path}',
                    'error_type': 'not_directory'
                }
            
            # プラットフォーム別にフォルダを開く
            try:
                if self.platform == "Windows":
                    os.startfile(folder_path)
                elif self.platform == "Darwin":  # macOS
                    subprocess.run(["open", folder_path], check=True)
                elif self.platform == "Linux":
                    subprocess.run(["xdg-open", folder_path], check=True)
                else:
                    return {
                        'success': False,
                        'error': f'サポートされていないプラットフォーム: {self.platform}',
                        'error_type': 'unsupported_platform',
                        'system_info': self.get_system_info()
                    }
            except subprocess.CalledProcessError as e:
                return {
                    'success': False,
                    'error': f'フォルダを開くコマンドが失敗しました: {str(e)}',
                    'error_type': 'command_failed'
                }
            except Exception as e:
                return {
                    'success': False,
                    'error': f'フォルダを開く際にエラーが発生しました: {str(e)}',
                    'error_type': 'open_failed'
                }
            
            return {
                'success': True,
                'message': f'フォルダを開きました: {folder_path}',
                'platform': self.platform
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'フォルダを開くエラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            } 