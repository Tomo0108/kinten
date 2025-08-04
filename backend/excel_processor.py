"""
Excel処理機能
テンプレートExcelにデータを転記する
"""

import openpyxl
import pandas as pd
from typing import Dict, Any, Optional
import os
from datetime import datetime


class ExcelProcessor:
    """Excel処理クラス"""
    
    def __init__(self):
        self.workbook = None  # type: ignore
        self.sheet = None  # type: ignore
    
    def load_template(self, template_path: str) -> Dict[str, Any]:
        """
        テンプレートExcelを読み込み
        
        Args:
            template_path: テンプレートファイルパス
            
        Returns:
            処理結果辞書
        """
        try:
            self.workbook = openpyxl.load_workbook(template_path)
            
            # 初期シート名「勤務表」を取得
            if "勤務表" in self.workbook.sheetnames:
                self.sheet = self.workbook["勤務表"]
            else:
                raise ValueError("テンプレートに「勤務表」シートが見つかりません")
            
            if self.sheet is None:
                return {
                    'success': False,
                    'error': "シートの取得に失敗しました"
                }
            
            return {
                'success': True,
                'sheet_name': self.sheet.title,
                'max_row': self.sheet.max_row,
                'max_column': self.sheet.max_column
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def update_sheet_name(self, year_month: str, employee_name: str = "") -> bool:
        """
        シート名を年月と氏名に応じて変更
        
        Args:
            year_month: 年月（例：202507）
            employee_name: 従業員名（オプション）
        """
        try:
            if self.sheet is None or self.workbook is None:
                print("シートまたはワークブックが初期化されていません")
                return False
                
            # 氏名が指定されている場合は含める
            if employee_name and employee_name.strip():
                new_sheet_name = f"勤怠表_{year_month}_{employee_name}"
            else:
                new_sheet_name = f"勤怠表_{year_month}"
            
            print(f"シート名を変更: {new_sheet_name}")
            
            # シート名変更
            self.sheet.title = new_sheet_name
            self.sheet = self.workbook[new_sheet_name]
            
            return True
            
        except Exception as e:
            print(f"シート名変更エラー: {e}")
            return False
    
    def write_employee_info(self, employee_name: str, year: str, month: str) -> bool:
        """
        従業員情報を書き込み
        
        Args:
            employee_name: 従業員名
            year: 年
            month: 月
        """
        try:
            if self.sheet is None:
                print("シートが初期化されていません")
                return False
                
            # G6: 従業員名
            self.sheet['G6'] = employee_name
            
            # F5: 年
            self.sheet['F5'] = year
            
            # H5: 月（先頭の0を除去して数値として設定）
            month_num = int(month)  # "06" → 6
            self.sheet['H5'] = month_num
            
            return True
            
        except Exception as e:
            print(f"従業員情報書き込みエラー: {e}")
            return False
    
    def write_attendance_data(self, df: pd.DataFrame, employee_name: str) -> bool:
        """
        勤怠データを転記
        
        Args:
            df: 勤怠データDataFrame
            employee_name: 従業員名（GUIから取得）
        """
        try:
            if self.sheet is None:
                print("シートが初期化されていません")
                return False
                
            # A11以降にデータを転記
            start_row = 11
            
            for index, row in df.iterrows():
                current_row = start_row + int(index)  # type: ignore
                
                # A列・B列: 日付と曜日は編集しない（雛形ファイルの関数が自動生成）
                # H5セルの月数を基準に雛形ファイルの関数が自動的に日付と曜日を設定
                
                # C列: 始業時刻
                self.sheet[f'C{current_row}'] = row['始業時刻1']
                
                # D列: 終業時刻
                self.sheet[f'D{current_row}'] = row['終業時刻1']
                
                # E列: 休憩時間（始業時刻と終業時刻がある場合は1:00）
                if pd.notna(row['始業時刻1']) and pd.notna(row['終業時刻1']) and \
                   str(row['始業時刻1']).strip() != '' and str(row['終業時刻1']).strip() != '':
                    self.sheet[f'E{current_row}'] = '1:00'
                else:
                    self.sheet[f'E{current_row}'] = ''
                
                # F列: 勤務時間（就業時間-始業時間-休憩時間）
                work_hours = self._calculate_work_hours(
                    row['始業時刻1'], 
                    row['終業時刻1'], 
                    '1:00'  # 固定の休憩時間
                )
                if work_hours > 0:
                    self.sheet[f'F{current_row}'] = work_hours
                else:
                    self.sheet[f'F{current_row}'] = ''
                
                # G列: 記入不要
                # 何も設定しない
                
                # H列: 詳細・備考（勤怠メモ）
                memo = row.get('勤怠メモ', '')
                if memo and str(memo).strip() != '':
                    self.sheet[f'H{current_row}'] = memo
            
            return True
            
        except Exception as e:
            print(f"勤怠データ転記エラー: {e}")
            return False
    
    def _calculate_work_hours(self, start_time, end_time, break_time):
        """
        勤務時間を計算（就業時間-始業時間-休憩時間）
        
        Args:
            start_time: 始業時刻
            end_time: 終業時刻
            break_time: 休憩時間
            
        Returns:
            勤務時間（時間単位、例：8.5）
        """
        try:
            # 空の値の場合は0を返す
            if pd.isna(start_time) or pd.isna(end_time) or \
               str(start_time).strip() == '' or str(end_time).strip() == '':
                return 0
            
            # 時刻を時間に変換
            def time_to_hours(time_str):
                if pd.isna(time_str) or str(time_str).strip() == '':
                    return 0
                
                time_str = str(time_str).strip()
                if ':' in time_str:
                    hours, minutes = map(int, time_str.split(':'))
                    return hours + minutes / 60.0
                else:
                    return 0
            
            start_hours = time_to_hours(start_time)
            end_hours = time_to_hours(end_time)
            break_hours = time_to_hours(break_time)
            
            # 勤務時間 = 終業時刻 - 始業時刻 - 休憩時間
            work_hours = end_hours - start_hours - break_hours
            
            # 負の値や0の場合は0を返す
            return max(0, round(work_hours, 1))
            
        except Exception as e:
            print(f"勤務時間計算エラー: {e}")
            return 0
    
    def save_workbook(self, output_path: str) -> Dict[str, Any]:
        """
        ワークブックを保存
        
        Args:
            output_path: 出力ファイルパス
            
        Returns:
            処理結果辞書
        """
        try:
            # 出力ディレクトリの存在確認と作成
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # ファイルが既に存在する場合の処理
            if os.path.exists(output_path):
                # ファイルが使用中かチェック
                try:
                    with open(output_path, 'r+b') as f:
                        pass
                except PermissionError:
                    # ファイルが使用中の場合は、タイムスタンプを付けて新しいファイル名を生成
                    base_name = os.path.splitext(output_path)[0]
                    extension = os.path.splitext(output_path)[1]
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_path = f"{base_name}_{timestamp}{extension}"
                    print(f"File is in use, saving as: {output_path}")
            
            # ファイル保存
            if self.workbook is None:
                return {
                    'success': False,
                    'error': "ワークブックが初期化されていません"
                }
            self.workbook.save(output_path)
            
            return {
                'success': True,
                'output_path': output_path
            }
            
        except PermissionError as e:
            return {
                'success': False,
                'error': f"権限エラー: ファイル '{output_path}' にアクセスできません。ファイルが他のアプリケーションで使用中か、権限が不足しています。"
            }
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Save workbook error: {error_details}")
            return {
                'success': False,
                'error': f"ファイル保存エラー: {str(e)}"
            } 