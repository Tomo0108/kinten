import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import os

def create_test_excel_files():
    """テスト用のExcelファイルを作成"""
    
    # テスト用ディレクトリを作成
    test_dir = "test_input"
    if not os.path.exists(test_dir):
        os.makedirs(test_dir)
    
    # テストファイル1: 勤怠表_202501_テスト1.xlsx
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    if ws1 is not None:
        ws1.title = "勤怠表_202501_テスト1"
        
        # ヘッダーを追加
        headers = ["日付", "曜日", "始業時刻1", "終業時刻1", "始業時刻2", "終業時刻2", "休憩時間", "勤務時間"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            if cell is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
    
    # サンプルデータを追加
    sample_data = [
        ["2025-01-01", "水", "09:00", "18:00", "", "", "60", "8.0"],
        ["2025-01-02", "木", "09:00", "17:30", "", "", "60", "7.5"],
        ["2025-01-03", "金", "08:30", "18:30", "", "", "60", "9.0"],
    ]
    
    if ws1 is not None:
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws1.cell(row=row, column=col, value=value)
    
    # ファイルを保存
    file1_path = os.path.join(test_dir, "勤怠表_202501_テスト1.xlsx")
    wb1.save(file1_path)
    print(f"✅ 作成完了: {file1_path}")
    
    # テストファイル2: 勤怠表_202501_テスト2.xlsx
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    if ws2 is not None:
        ws2.title = "勤怠表_202501_テスト2"
        
        # ヘッダーを追加
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            if cell is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        
        # サンプルデータを追加
        sample_data2 = [
            ["2025-01-06", "月", "09:00", "18:00", "", "", "60", "8.0"],
            ["2025-01-07", "火", "09:00", "17:30", "", "", "60", "7.5"],
            ["2025-01-08", "水", "08:30", "18:30", "", "", "60", "9.0"],
        ]
        
        for row, data in enumerate(sample_data2, 2):
            for col, value in enumerate(data, 1):
                ws2.cell(row=row, column=col, value=value)
    
    # ファイルを保存
    file2_path = os.path.join(test_dir, "勤怠表_202501_テスト2.xlsx")
    wb2.save(file2_path)
    print(f"✅ 作成完了: {file2_path}")
    
    # テストファイル3: 勤怠表_202501_テスト3.xlsx
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    if ws3 is not None:
        ws3.title = "勤怠表_202501_テスト3"
        
        # ヘッダーを追加
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            if cell is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        
        # サンプルデータを追加
        sample_data3 = [
            ["2025-01-09", "木", "09:00", "18:00", "", "", "60", "8.0"],
            ["2025-01-10", "金", "09:00", "17:30", "", "", "60", "7.5"],
            ["2025-01-13", "月", "08:30", "18:30", "", "", "60", "9.0"],
        ]
        
        for row, data in enumerate(sample_data3, 2):
            for col, value in enumerate(data, 1):
                ws3.cell(row=row, column=col, value=value)
    
    # ファイルを保存
    file3_path = os.path.join(test_dir, "勤怠表_202501_テスト3.xlsx")
    wb3.save(file3_path)
    print(f"✅ 作成完了: {file3_path}")
    
    print(f"\n📁 テスト用Excelファイルが {test_dir} ディレクトリに作成されました")
    return [file1_path, file2_path, file3_path]

if __name__ == "__main__":
    create_test_excel_files() 