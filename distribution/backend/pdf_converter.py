#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF変換機能（クロスプラットフォーム対応）
ExcelファイルをPDFに変換する - Windows/Mac対応
"""

import os
import glob
import platform
import subprocess
import sys
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path

# Excel読み込み用
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl not available")
    OPENPYXL_AVAILABLE = False
    # フォールバック用のダミー定義
    class Worksheet:
        pass

# PDF作成用
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.pdfbase import pdfutils
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    REPORTLAB_AVAILABLE = True
except ImportError:
    print("Warning: reportlab not available")
    REPORTLAB_AVAILABLE = False

# 日本語フォント対応
JAPANESE_FONT = 'Helvetica'
if REPORTLAB_AVAILABLE:
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        JAPANESE_FONT = 'HeiseiMin-W3'
    except:
        # フォントが利用できない場合はHelveticaを使用
        JAPANESE_FONT = 'Helvetica'


class PDFConverter:
    """クロスプラットフォーム対応PDF変換クラス"""
    
    def __init__(self):
        self.platform = platform.system()
        self._check_dependencies()
        
    def _check_dependencies(self) -> Dict[str, bool]:
        """依存関係の確認"""
        deps = {
            'openpyxl': OPENPYXL_AVAILABLE,
            'reportlab': REPORTLAB_AVAILABLE,
            'platform': True
        }
        return deps
    
    def get_system_info(self) -> Dict[str, Any]:
        """システム情報を取得"""
        deps = self._check_dependencies()
        return {
            'platform': self.platform,
            'python_version': sys.version,
            'dependencies': deps,
            'missing_deps': [k for k, v in deps.items() if not v]
        }
        
    def get_excel_files(self, folder_path: str) -> Dict[str, Any]:
        """
        指定フォルダ内のExcelファイルを取得
        
        Args:
            folder_path: フォルダパス
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            folder_path = os.path.normpath(folder_path)
            
            if not os.path.exists(folder_path):
                return {
                    'success': False,
                    'error': f'フォルダが見つかりません: {folder_path}',
                    'error_type': 'folder_not_found'
                }
            
            if not os.path.isdir(folder_path):
                return {
                    'success': False,
                    'error': f'指定されたパスはフォルダではありません: {folder_path}',
                    'error_type': 'not_directory'
                }
            
            # Excelファイルを検索
            excel_patterns = ['*.xlsx', '*.xls', '*.xlsm']
            excel_files = []
            
            for pattern in excel_patterns:
                try:
                    files = glob.glob(os.path.join(folder_path, pattern))
                    excel_files.extend(files)
                except Exception as e:
                    print(f"Warning: Failed to search pattern {pattern}: {str(e)}")
                    continue
            
            # ファイル情報を取得
            file_list = []
            failed_files = []
            
            for file_path in excel_files:
                try:
                    file_name = os.path.basename(file_path)
                    file_size = os.path.getsize(file_path)
                    
                    # ファイルの読み取り権限を確認
                    if not os.access(file_path, os.R_OK):
                        failed_files.append({
                            'path': file_path,
                            'error': 'ファイルの読み取り権限がありません'
                        })
                        continue
                    
                    file_list.append({
                        'path': file_path,
                        'name': file_name,
                        'size': file_size
                    })
                except Exception as e:
                    failed_files.append({
                        'path': file_path,
                        'error': f"ファイル情報取得エラー: {str(e)}"
                    })
                    continue
            
            result = {
                'success': True,
                'files': file_list,
                'count': len(file_list),
                'folder_path': folder_path
            }
            
            if failed_files:
                result['failed_files'] = failed_files
                result['warning'] = f'{len(failed_files)}個のファイルでエラーが発生しました'
            
            return result
            
        except Exception as e:
            return {
                'success': False,
                'error': f'ファイル取得エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def create_output_folder(self, base_output_dir: str) -> Dict[str, Any]:
        """
        出力フォルダを作成（日付フォルダ）
        
        Args:
            base_output_dir: 基本出力ディレクトリ
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            base_output_dir = os.path.normpath(base_output_dir)
            
            # 親ディレクトリの存在確認
            if not os.path.exists(base_output_dir):
                try:
                    Path(base_output_dir).mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    return {
                        'success': False,
                        'error': f'ベースディレクトリの作成に失敗: {str(e)}',
                        'error_type': 'base_dir_creation_failed'
                    }
            
            # 現在の日付でフォルダ名を生成（YYYYMM形式）
            current_date = datetime.now()
            folder_name = current_date.strftime("%Y%m")
            
            # 出力フォルダパスを作成
            output_folder = os.path.join(base_output_dir, folder_name)
            
            # フォルダを作成
            try:
                Path(output_folder).mkdir(parents=True, exist_ok=True)
            except PermissionError:
                return {
                    'success': False,
                    'error': f'フォルダの作成権限がありません: {output_folder}',
                    'error_type': 'permission_denied'
                }
            except Exception as e:
                return {
                    'success': False,
                    'error': f'フォルダ作成エラー: {str(e)}',
                    'error_type': 'folder_creation_failed'
                }
            
            # 書き込み権限の確認
            if not os.access(output_folder, os.W_OK):
                return {
                    'success': False,
                    'error': f'出力フォルダに書き込み権限がありません: {output_folder}',
                    'error_type': 'write_permission_denied'
                }
            
            return {
                'success': True,
                'output_folder': output_folder,
                'folder_name': folder_name,
                'base_dir': base_output_dir
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'出力フォルダ作成エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def _validate_excel_file(self, file_path: str) -> Tuple[bool, str]:
        """Excelファイルの妥当性を検証"""
        try:
            if not OPENPYXL_AVAILABLE:
                return False, "openpyxl が利用できません"
            
            if not os.path.exists(file_path):
                return False, "ファイルが見つかりません"
            
            if not os.access(file_path, os.R_OK):
                return False, "ファイルの読み取り権限がありません"
            
            # ファイルサイズの確認
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                return False, "ファイルが空です"
            
            if file_size > 100 * 1024 * 1024:  # 100MB制限
                return False, "ファイルサイズが大きすぎます (100MB以上)"
            
            # Excelファイルの読み込みテスト
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                if not workbook.sheetnames:
                    return False, "シートが見つかりません"
                workbook.close()
            except Exception as e:
                return False, f"Excelファイルの読み込みエラー: {str(e)}"
            
            return True, "OK"
            
        except Exception as e:
            return False, f"ファイル検証エラー: {str(e)}"
    
    def _excel_to_pdf_openpyxl(self, excel_file: str, pdf_path: str) -> Tuple[bool, str]:
        """
        openpyxlを使用してExcelからPDFを生成
        
        Args:
            excel_file: Excelファイルのパス
            pdf_path: 出力PDFのパス
            
        Returns:
            (成功/失敗, エラーメッセージ)
        """
        try:
            if not REPORTLAB_AVAILABLE:
                return False, "reportlab が利用できません"
            
            # Excelファイルを読み込み
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # PDFドキュメントを作成
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=20*mm,
                bottomMargin=20*mm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # タイトルスタイル
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=JAPANESE_FONT,
                fontSize=16,
                spaceAfter=12
            )
            
            # ファイル名をタイトルとして追加
            file_title = os.path.splitext(os.path.basename(excel_file))[0]
            story.append(Paragraph(f"Excel File: {file_title}", title_style))
            story.append(Spacer(1, 12))
            
            # 各シートを処理
            processed_sheets = 0
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    
                    # シート名をタイトルとして追加
                    story.append(Paragraph(f"シート: {sheet_name}", title_style))
                    story.append(Spacer(1, 12))
                    
                    # シートのデータを取得
                    data = self._get_sheet_data(sheet)
                    
                    if data:
                        # テーブルを作成
                        table = Table(data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('FONTNAME', (0, 1), (-1, -1), JAPANESE_FONT),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        story.append(table)
                        processed_sheets += 1
                    else:
                        story.append(Paragraph("データが見つかりません", styles['Normal']))
                    
                    story.append(PageBreak())
                    
                except Exception as e:
                    error_msg = f"シート '{sheet_name}' の処理エラー: {str(e)}"
                    story.append(Paragraph(error_msg, styles['Normal']))
                    story.append(PageBreak())
                    continue
            
            workbook.close()
            
            if processed_sheets == 0:
                return False, "処理可能なシートが見つかりませんでした"
            
            # PDFを生成
            doc.build(story)
            
            # ファイルが正常に作成されたかチェック
            if not os.path.exists(pdf_path):
                return False, "PDFファイルの作成に失敗しました"
            
            if os.path.getsize(pdf_path) == 0:
                return False, "PDFファイルが空です"
            
            return True, f"成功 ({processed_sheets} シート処理)"
            
        except Exception as e:
            return False, f"PDF変換エラー: {str(e)}"
    
    def _get_sheet_data(self, sheet: Worksheet, max_rows: int = 100, max_cols: int = 20) -> List[List[str]]:
        """
        シートからデータを取得
        
        Args:
            sheet: ワークシート
            max_rows: 最大行数
            max_cols: 最大列数
            
        Returns:
            データリスト
        """
        try:
            data = []
            
            # 実際に使用されている範囲を取得
            max_row = min(sheet.max_row, max_rows) if sheet.max_row else 1
            max_col = min(sheet.max_column, max_cols) if sheet.max_column else 1
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        value = cell.value
                        
                        # 値を文字列に変換
                        if value is None:
                            row_data.append('')
                        elif isinstance(value, (int, float)):
                            row_data.append(str(value))
                        else:
                            # 文字列として扱い、長すぎる場合は切り詰め
                            str_value = str(value)
                            if len(str_value) > 100:
                                str_value = str_value[:100] + "..."
                            row_data.append(str_value)
                    except Exception:
                        row_data.append('')
                
                # 空行をスキップ（すべてのセルが空の場合）
                if any(cell.strip() for cell in row_data if isinstance(cell, str)):
                    data.append(row_data)
            
            return data
            
        except Exception as e:
            print(f"シートデータ取得エラー: {str(e)}")
            return []
    
    def convert_to_pdf(self, excel_files: List[str], output_folder: str) -> Dict[str, Any]:
        """
        ExcelファイルをPDFに変換（クロスプラットフォーム対応）
        
        Args:
            excel_files: 変換するExcelファイルのパスリスト
            output_folder: 出力フォルダパス
            
        Returns:
            結果辞書
        """
        try:
            # 依存関係の確認
            deps = self._check_dependencies()
            missing_deps = [k for k, v in deps.items() if not v]
            
            if missing_deps:
                return {
                    'success': False,
                    'error': f'必要なライブラリが不足しています: {", ".join(missing_deps)}',
                    'error_type': 'missing_dependencies',
                    'missing_dependencies': missing_deps,
                    'system_info': self.get_system_info()
                }
            
            # 出力フォルダの存在確認
            if not os.path.exists(output_folder):
                return {
                    'success': False,
                    'error': f'出力フォルダが見つかりません: {output_folder}',
                    'error_type': 'output_folder_not_found'
                }
            
            converted_files = []
            failed_files = []
            validation_errors = []
            
            for excel_file in excel_files:
                try:
                    # ファイルの妥当性を検証
                    is_valid, validation_message = self._validate_excel_file(excel_file)
                    if not is_valid:
                        validation_errors.append({
                            'file': excel_file,
                            'error': validation_message
                        })
                        continue
                    
                    # ファイル名からPDF名を生成
                    base_name = os.path.splitext(os.path.basename(excel_file))[0]
                    pdf_name = f"{base_name}.pdf"
                    pdf_path = os.path.join(output_folder, pdf_name)
                    
                    # 既存ファイルの確認
                    if os.path.exists(pdf_path):
                        # タイムスタンプを追加
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        pdf_name = f"{base_name}_{timestamp}.pdf"
                        pdf_path = os.path.join(output_folder, pdf_name)
                    
                    # PDFに変換
                    success, message = self._excel_to_pdf_openpyxl(excel_file, pdf_path)
                    
                    if success and os.path.exists(pdf_path):
                        converted_files.append({
                            'excel_file': excel_file,
                            'pdf_file': pdf_path,
                            'pdf_name': pdf_name,
                            'message': message
                        })
                    else:
                        failed_files.append({
                            'file': excel_file,
                            'error': message or 'PDF変換に失敗しました'
                        })
                    
                except Exception as e:
                    failed_files.append({
                        'file': excel_file,
                        'error': f'予期しないエラー: {str(e)}'
                    })
            
            # 結果の整理
            result = {
                'success': True,
                'converted_files': converted_files,
                'failed_files': failed_files,
                'validation_errors': validation_errors,
                'total_converted': len(converted_files),
                'total_failed': len(failed_files),
                'total_validation_errors': len(validation_errors),
                'output_folder': output_folder
            }
            
            # 全て失敗した場合
            if len(converted_files) == 0 and (len(failed_files) > 0 or len(validation_errors) > 0):
                result['success'] = False
                result['error'] = 'すべてのファイルの変換に失敗しました'
            
            return result
            
        except Exception as e:
            return {
                'success': False,
                'error': f'PDF変換エラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            }
    
    def open_folder(self, folder_path: str) -> Dict[str, Any]:
        """
        フォルダを開く（クロスプラットフォーム対応）
        
        Args:
            folder_path: 開くフォルダのパス
            
        Returns:
            結果辞書
        """
        try:
            # パスの正規化
            folder_path = os.path.normpath(folder_path)
            
            if not os.path.exists(folder_path):
                return {
                    'success': False,
                    'error': f'フォルダが見つかりません: {folder_path}',
                    'error_type': 'folder_not_found'
                }
            
            if not os.path.isdir(folder_path):
                return {
                    'success': False,
                    'error': f'指定されたパスはフォルダではありません: {folder_path}',
                    'error_type': 'not_directory'
                }
            
            # プラットフォーム別にフォルダを開く
            try:
                if self.platform == "Windows":
                    os.startfile(folder_path)
                elif self.platform == "Darwin":  # macOS
                    subprocess.run(["open", folder_path], check=True)
                elif self.platform == "Linux":
                    subprocess.run(["xdg-open", folder_path], check=True)
                else:
                    return {
                        'success': False,
                        'error': f'サポートされていないプラットフォーム: {self.platform}',
                        'error_type': 'unsupported_platform',
                        'system_info': self.get_system_info()
                    }
            except subprocess.CalledProcessError as e:
                return {
                    'success': False,
                    'error': f'フォルダを開くコマンドが失敗しました: {str(e)}',
                    'error_type': 'command_failed'
                }
            except Exception as e:
                return {
                    'success': False,
                    'error': f'フォルダを開く際にエラーが発生しました: {str(e)}',
                    'error_type': 'open_failed'
                }
            
            return {
                'success': True,
                'message': f'フォルダを開きました: {folder_path}',
                'platform': self.platform
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'フォルダを開くエラー: {str(e)}',
                'error_type': 'general_error',
                'system_info': self.get_system_info()
            } 