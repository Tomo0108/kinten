"""
サンプルExcelテンプレート作成スクリプト
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_sample_template():
    """サンプル勤怠表テンプレートを作成"""
    
    # ワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "勤務表"
    
    # スタイル定義
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # ヘッダー部分
    ws['A1'] = '勤怠表'
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = center_alignment
    
    # 年月表示
    ws['F5'] = '年'
    ws['H5'] = '月'
    ws['F5'].font = header_font
    ws['H5'].font = header_font
    ws['F5'].alignment = center_alignment
    ws['H5'].alignment = center_alignment
    
    # 従業員名
    ws['G6'] = '従業員名'
    ws['G6'].font = header_font
    ws['G6'].alignment = center_alignment
    
    # テーブルヘッダー
    headers = ['日付', '曜日', '始業時刻', '終業時刻', '休憩時間', '総勤務時間', '備考']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # データ行のサンプル（1-31日）
    for day in range(1, 32):
        row = 10 + day
        ws.cell(row=row, column=1, value=f'2025-07-{day:02d}')  # 日付
        ws.cell(row=row, column=2, value='')  # 曜日
        ws.cell(row=row, column=3, value='')  # 始業時刻
        ws.cell(row=row, column=4, value='')  # 終業時刻
        ws.cell(row=row, column=5, value='')  # 休憩時間
        ws.cell(row=row, column=6, value='')  # 総勤務時間
        ws.cell(row=row, column=7, value='')  # 備考
        
        # ボーダーを適用
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
    
    # 列幅調整
    column_widths = [12, 8, 10, 10, 10, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # ファイル保存
    output_path = os.path.join('assets', '勤怠表_ひな型.xlsx')
    wb.save(output_path)
    print(f"サンプルテンプレートを作成しました: {output_path}")
    
    return output_path

if __name__ == "__main__":
    create_sample_template() 